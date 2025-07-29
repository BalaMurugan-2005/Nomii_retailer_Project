from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from flask.json.provider import JSONProvider
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import openpyxl
from datetime import datetime
import os
import speech_recognition as sr
import re
from fpdf import FPDF
import random
from collections import defaultdict
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import numpy as np
import os

from json import JSONEncoder

class CustomJSONEncoder(JSONProvider):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int64)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        return super().default(obj)

app = Flask(__name__)
app.json_encoder = CustomJSONEncoder
app.secret_key = 'your_secret_key_here'

# Data file paths
PRODUCTS_FILE = 'data/Products.xlsx'
ORDERS_FILE = 'data/retailer_orders.xlsx'
AI_SUGGESTIONS_FILE = 'data/Ai_suggestion_products.xlsx'
DELIVERY_STATUS_FILE = 'data/deliverystatus.xlsx'
MONEY_SPENT_FILE = 'data/MoneySpent.xlsx'
REWARDS_FILE = 'data/retailer_rewards.xlsx'
USERS_FILE = 'data/retailer_users.xlsx'

# Initialize data files if they don't exist
if not os.path.exists('data'):
    os.makedirs('data')

required_files = {
    PRODUCTS_FILE: ['ProductID', 'Name', 'Category', 'Price', 'Supplier', 'Stock'],
    ORDERS_FILE: ['OrderID', 'ProductID', 'ProductName', 'Quantity', 'Price', 'Total', 'OrderDate', 'Status'],
    AI_SUGGESTIONS_FILE: ['ProductID', 'Name', 'Category', 'Reason'],
    DELIVERY_STATUS_FILE: ['OrderID', 'Status', 'LastUpdate', 'DeliveryAgent'],
    MONEY_SPENT_FILE: ['TransactionID', 'Amount', 'Date', 'Description'],
    REWARDS_FILE: ['Points', 'Badges', 'Level'],
    USERS_FILE: ['ShopName', 'OwnerName', 'Location', 'Phone', 'Email', 'Password']
}

for file, columns in required_files.items():
    if not os.path.exists(file):
        df = pd.DataFrame(columns=columns)
        df.to_excel(file, index=False)

# Helper functions
def get_next_id(filename, id_column):
    try:
        df = pd.read_excel(filename)
        if df.empty:
            return 1
        return int(df[id_column].max()) + 1  # Ensure native Python int
    except:
        return 1

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

def save_to_excel(data, filename):
    try:
        df = pd.read_excel(filename)
        new_df = pd.DataFrame([data])
        df = pd.concat([df, new_df], ignore_index=True)
        df.to_excel(filename, index=False)
        return True
    except Exception as e:
        print(f"Error saving to {filename}: {e}")
        return False

def get_user_orders():
    try:
        df = pd.read_excel(ORDERS_FILE)
        # Convert numeric columns to native Python types
        numeric_cols = ['ProductID', 'Quantity', 'Price', 'Total']
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric)
        return df.to_dict('records')
    except Exception as e:
        print(f"Error getting user orders: {e}")
        return []

def get_product_suggestions():
    try:
        df = pd.read_excel(AI_SUGGESTIONS_FILE)
        # Convert numeric columns to native Python types
        if 'ProductID' in df.columns:
            df['ProductID'] = df['ProductID'].astype(int)
        return df.sample(5).to_dict('records')
    except Exception as e:
        print(f"Error getting product suggestions: {e}")
        return []

def generate_restock_predictions():
    try:
        orders_df = pd.read_excel(ORDERS_FILE)
        
        if orders_df.empty:
            return []
            
        freq_products = orders_df['ProductName'].value_counts().head(5).index.tolist()
        recent_orders = orders_df.sort_values('OrderDate', ascending=False)
        
        predictions = []
        for product in freq_products:
            last_ordered = recent_orders[recent_orders['ProductName'] == product]
            if not last_ordered.empty:
                days_since = (datetime.now() - last_ordered.iloc[0]['OrderDate']).days
                if days_since > 7:
                    predictions.append({
                        'product': product,
                        'message': f"Restock soon! You usually order this every {random.randint(3,5)} days",
                        'urgency': 'high' if days_since > 14 else 'medium'
                    })
        
        return predictions[:3]
    except Exception as e:
        print(f"Error generating restock predictions: {e}")
        return []

def generate_combo_suggestions():
    try:
        orders_df = pd.read_excel(ORDERS_FILE)
        
        if orders_df.empty:
            return []
            
        order_ids = orders_df['OrderID'].unique()
        product_pairs = defaultdict(int)
        
        for order_id in order_ids:
            products = orders_df[orders_df['OrderID'] == order_id]['ProductName'].tolist()
            for i in range(len(products)):
                for j in range(i+1, len(products)):
                    pair = tuple(sorted([products[i], products[j]]))
                    product_pairs[pair] += 1
        
        top_pairs = sorted(product_pairs.items(), key=lambda x: x[1], reverse=True)[:3]
        
        suggestions = []
        for pair, count in top_pairs:
            discount = random.randint(5, 15)
            suggestions.append({
                'products': f"{pair[0]} + {pair[1]}",
                'discount': f"₹{discount} off",
                'reason': f"Frequently bought together ({count} times)"
            })
        
        return suggestions
    except Exception as e:
        print(f"Error generating combo suggestions: {e}")
        return []

def generate_weekly_insights():
    try:
        orders_df = pd.read_excel(ORDERS_FILE)
        
        if orders_df.empty:
            return {}
            
        orders_df['OrderDate'] = pd.to_datetime(orders_df['OrderDate'])
        weekly_spending = orders_df.groupby(
            orders_df['OrderDate'].dt.to_period('W'))['Total'].sum().reset_index()
        weekly_spending['OrderDate'] = weekly_spending['OrderDate'].astype(str)
        
        top_products = orders_df.groupby('ProductName')['Quantity'].sum().nlargest(5)
        
        avg_order_value = float(orders_df['Total'].mean())  # Convert to native float
        order_count = int(len(orders_df['OrderID'].unique()))  # Convert to native int
        
        return {
            'weekly_spending': weekly_spending.to_dict('records'),
            'top_products': top_products.reset_index().to_dict('records'),
            'avg_order_value': round(avg_order_value, 2),
            'order_count': order_count
        }
    except Exception as e:
        print(f"Error generating weekly insights: {e}")
        return {}

def generate_pdf_invoice(order_data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    pdf.cell(200, 10, txt="Nomii - Invoice", ln=1, align='C')
    pdf.cell(200, 10, txt=f"Order ID: {order_data['OrderID']}", ln=1, align='L')
    pdf.cell(200, 10, txt=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='L')
    
    pdf.cell(200, 10, txt=f"Retailer: {session.get('shop_name')}", ln=1, align='L')
    pdf.cell(200, 10, txt=f"Location: {session.get('location')}", ln=1, align='L')
    
    pdf.cell(200, 10, txt="Items:", ln=1, align='L')
    pdf.cell(100, 10, txt="Product", border=1)
    pdf.cell(30, 10, txt="Qty", border=1)
    pdf.cell(30, 10, txt="Price", border=1)
    pdf.cell(30, 10, txt="Total", border=1, ln=1)
    
    for item in order_data['items']:
        pdf.cell(100, 10, txt=item['ProductName'], border=1)
        pdf.cell(30, 10, txt=str(item['Quantity']), border=1)
        pdf.cell(30, 10, txt=str(item['Price']), border=1)
        pdf.cell(30, 10, txt=str(item['Total']), border=1, ln=1)
    
    pdf.cell(160, 10, txt="Total Amount:", border=1)
    pdf.cell(30, 10, txt=str(order_data['total_amount']), border=1, ln=1)
    
    buffer = BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer

# Routes
@app.route('/')
def home():
    if 'email' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        try:
            df = pd.read_excel(USERS_FILE)
            user = df[df['Email'] == email].iloc[0]
            
            if user is not None and check_password_hash(user['Password'], password):
                session['shop_name'] = user['ShopName']
                session['location'] = user['Location']
                session['email'] = user['Email']
                return redirect(url_for('dashboard'))
            else:
                return render_template('login.html', error="Invalid credentials")
        except Exception as e:
            print(f"Login error: {e}")
            return render_template('login.html', error="User not found")
    
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        try:
            df = pd.read_excel(USERS_FILE)
            
            if not df[df['Email'] == request.form['email']].empty:
                return render_template('signup.html', error="Email already registered")
            
            new_user = {
                'ShopName': request.form['shop_name'],
                'OwnerName': request.form['owner_name'],
                'Location': request.form['location'],
                'Phone': request.form['phone'],
                'Email': request.form['email'],
                'Password': generate_password_hash(request.form['password'])
            }
            
            save_to_excel(new_user, USERS_FILE)
            return redirect(url_for('login'))
        except Exception as e:
            print(f"Signup error: {e}")
            return render_template('signup.html', error=f"Registration failed: {str(e)}")
    
    return render_template('signup.html')

@app.route('/dashboard')
def dashboard():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    orders = get_user_orders()
    suggestions = get_product_suggestions()
    restock_predictions = generate_restock_predictions()
    combo_suggestions = generate_combo_suggestions()
    weekly_insights = generate_weekly_insights()
    
    recent_orders = sorted(orders, key=lambda x: x['OrderDate'], reverse=True)[:5] if orders else []
    
    try:
        delivery_df = pd.read_excel(DELIVERY_STATUS_FILE)
        delivery_statuses = delivery_df[delivery_df['OrderID'].isin([o['OrderID'] for o in orders])].to_dict('records')
    except Exception as e:
        print(f"Error getting delivery statuses: {e}")
        delivery_statuses = []
    
    return render_template('dashboard.html', 
                         recent_orders=recent_orders,
                         suggestions=suggestions,
                         restock_predictions=restock_predictions,
                         combo_suggestions=combo_suggestions,
                         weekly_insights=weekly_insights,
                         delivery_statuses=delivery_statuses)

@app.route('/products')
def products():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    search_query = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    
    try:
        df = pd.read_excel(PRODUCTS_FILE)
        # Remove the int conversion since ProductID has letters
        # df['ProductID'] = df['ProductID'].astype(int)  # Remove this line
        df['Price'] = df['Price'].astype(float)
        
        if search_query:
            df = df[df['Name'].str.contains(search_query, case=False)]
        if category_filter:
            df = df[df['Category'] == category_filter]
            
        categories = df['Category'].unique().tolist()
        products = df.to_dict('records')
    except Exception as e:
        print(f"Error getting products: {e}")
        flash("Error loading products. Please try again.", "danger")
        categories = []
        products = []
    
    return render_template('products.html', 
                         products=products, 
                         categories=categories,
                         search_query=search_query,
                         selected_category=category_filter)

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    if 'email' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    product_id = request.form.get('product_id')  # This will be in format "P001", "P002", etc.
    quantity = int(request.form.get('quantity', 1))
    
    try:
        df = pd.read_excel(PRODUCTS_FILE)
        # Match the ProductID as string without any conversion
        product = df[df['ProductID'] == product_id].iloc[0]
        
        cart = session.get('cart', [])
        
        # Check if product already exists in cart
        found = False
        for item in cart:
            if item['ProductID'] != product_id:
                item['Quantity'] += quantity
                item['Total'] = round(float(item['Quantity'] * float(item['Price'])), 2)
                found = True
                break
        
        # If not found, add new item to cart
        if not found:
            cart.append({
                 # Keep as string
                'ProductName': product['Name'],
                'Quantity': quantity,
                'Price': round(float(product['Price']), 2),
                'Total': round(float(quantity * float(product['Price'])), 2)
            })
        
        session['cart'] = cart
        return jsonify({
            'success': True, 
            'cart_size': len(cart),
            'message': f"{product['Name']} added to cart"
        })
    except IndexError:
        return jsonify({'success': False, 'error': 'Product not found'})
    except ValueError as ve:
        return jsonify({'success': False, 'error': f'Invalid quantity: {str(ve)}'})
    except Exception as e:
        print(f"Error adding to cart: {e}")
        return jsonify({'success': False, 'error': 'Failed to add product to cart'})
@app.route('/cart')
def view_cart():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    cart = session.get('cart', [])
    total = sum(float(item['Total']) for item in cart)  # Ensure float
    
    return render_template('cart.html', cart=cart, total=total)

@app.route('/update_cart', methods=['POST'])
def update_cart():
    if 'email' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    product_id = request.form.get('product_id')
    quantity = int(request.form.get('quantity', 1))
    
    cart = session.get('cart', [])
    
    for item in cart:
        if item['ProductID'] == product_id:
            if quantity <= 0:
                cart.remove(item)
            else:
                item['Quantity'] = quantity
                item['Total'] = float(quantity * float(item['Price']))  # Ensure float
            break
    
    session['cart'] = cart
    return jsonify({'success': True, 'cart_size': len(cart)})

@app.route('/place_order', methods=['POST'])
def place_order():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    cart = session.get('cart', [])
    if not cart:
        return redirect(url_for('view_cart'))
    
    order_id = get_next_id(ORDERS_FILE, 'OrderID')
    order_date = datetime.now()
    
    try:
        for item in cart:
            order_data = {
                'OrderID': order_id,
                'ProductID': int(item['ProductID']),  # Ensure native int
                'ProductName': item['ProductName'],
                'Quantity': int(item['Quantity']),  # Ensure native int
                'Price': float(item['Price']),  # Ensure native float
                'Total': float(item['Total']),  # Ensure native float
                'OrderDate': order_date,
                'Status': 'Ordered'
            }
            save_to_excel(order_data, ORDERS_FILE)
        
        total_amount = sum(float(item['Total']) for item in cart)  # Ensure float
        transaction_data = {
            'TransactionID': get_next_id(MONEY_SPENT_FILE, 'TransactionID'),
            'Amount': float(total_amount),  # Ensure native float
            'Date': order_date,
            'Description': f"Order #{order_id}"
        }
        save_to_excel(transaction_data, MONEY_SPENT_FILE)
        
        delivery_status = {
            'OrderID': order_id,
            'Status': 'Ordered',
            'LastUpdate': order_date,
            'DeliveryAgent': f"Agent {random.randint(1000, 9999)}"
        }
        save_to_excel(delivery_status, DELIVERY_STATUS_FILE)
        
        update_rewards(total_amount)
        
        session.pop('cart', None)
        
        invoice_data = {
            'OrderID': order_id,
            'items': cart,
            'total_amount': total_amount,
            'order_date': order_date
        }
        
        return render_template('order_success.html', order_id=order_id, total=total_amount)
    except Exception as e:
        print(f"Error placing order: {e}")
        return render_template('cart.html', error=f"Order failed: {str(e)}")

def update_rewards(amount):
    try:
        df = pd.read_excel(REWARDS_FILE)
        
        points_earned = int(float(amount) / 10)  # Ensure proper calculation
        
        if df.empty:
            new_rewards = {
                'Points': points_earned,
                'Badges': 'Newbie',
                'Level': 1
            }
            save_to_excel(new_rewards, REWARDS_FILE)
        else:
            df.at[0, 'Points'] = int(df.at[0, 'Points']) + points_earned  # Ensure native int
            
            if int(df.at[0, 'Points']) >= 100 and int(df.at[0, 'Level']) == 1:
                df.at[0, 'Level'] = 2
                df.at[0, 'Badges'] = 'Bronze'
            elif int(df.at[0, 'Points']) >= 500 and int(df.at[0, 'Level']) == 2:
                df.at[0, 'Level'] = 3
                df.at[0, 'Badges'] = 'Silver'
            elif int(df.at[0, 'Points']) >= 1000 and int(df.at[0, 'Level']) == 3:
                df.at[0, 'Level'] = 4
                df.at[0, 'Badges'] = 'Gold'
            
            df.to_excel(REWARDS_FILE, index=False)
    except Exception as e:
        print(f"Error updating rewards: {e}")

@app.route('/orders')
def orders():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    try:
        # Read all orders without any filtering
        orders_df = pd.read_excel(ORDERS_FILE)
        
        # Convert to list of dictionaries
        all_orders = orders_df.to_dict('records')

        return render_template('orders.html', 
                            orders=all_orders,
                            status_filter='',
                            date_from='',
                            date_to='')
    
    except Exception as e:
        print(f"Error loading orders: {e}")
        return render_template('orders.html', 
                            orders=[],
                            status_filter='',
                            date_from='',
                            date_to='')

@app.route('/download_invoice/<int:order_id>')
def download_invoice(order_id):
    if 'email' not in session:
        return redirect(url_for('login'))
    
    try:
        df = pd.read_excel(ORDERS_FILE)
        order_items = df[df['OrderID'] == order_id].to_dict('records')
        
        if not order_items:
            return "Order not found", 404
        
        total_amount = sum(float(item['Total']) for item in order_items)  # Ensure float
        order_date = order_items[0]['OrderDate']
        
        invoice_data = {
            'OrderID': order_id,
            'items': order_items,
            'total_amount': total_amount,
            'order_date': order_date
        }
        
        pdf_buffer = generate_pdf_invoice(invoice_data)
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"invoice_{order_id}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        print(f"Error generating invoice: {e}")
        return f"Error generating invoice: {str(e)}", 500

@app.route('/voice_order', methods=['POST'])
def voice_order():
    if 'email' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    try:
        audio_file = request.files['audio']
        r = sr.Recognizer()
        
        with sr.AudioFile(audio_file) as source:
            audio_data = r.record(source)
            text = r.recognize_google(audio_data)
            
        pattern = r'(?:add|order)\s+(\d+)\s*(kg|g|ml|l)?\s+(.+)'
        match = re.search(pattern, text, re.IGNORECASE)
        
        if not match:
            return jsonify({'success': False, 'error': 'Could not understand command'})
        
        quantity = int(match.group(1))
        unit = match.group(2) or ''
        product_name = match.group(3).strip()
        
        df = pd.read_excel(PRODUCTS_FILE)
        product = df[df['Name'].str.contains(product_name, case=False)].iloc[0]
        
        cart = session.get('cart', [])
        
        found = False
        for item in cart:
            if item['ProductID'] == product['ProductID']:
                item['Quantity'] += quantity
                item['Total'] = float(item['Quantity'] * float(item['Price']))  # Ensure float
                found = True
                break
        
        if not found:
            cart.append({
                'ProductID': int(product['ProductID']),  # Ensure native int
                'ProductName': product['Name'],
                'Quantity': quantity,
                'Price': float(product['Price']),  # Ensure native float
                'Total': float(quantity * float(product['Price']))  # Ensure float
            })
        
        session['cart'] = cart
        return jsonify({
            'success': True,
            'message': f"Added {quantity}{unit} {product_name} to cart",
            'cart_size': len(cart)
        })
    except Exception as e:
        print(f"Error processing voice order: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/ai_assistant', methods=['GET', 'POST'])
def ai_assistant():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        query = request.form.get('query', '').lower()
        response = ""
        
        if 'track' in query and 'order' in query:
            last_order = sorted(get_user_orders(), key=lambda x: x['OrderDate'], reverse=True)[0]
            response = f"Your last order #{last_order['OrderID']} is {last_order['Status']}"
        elif 'suggest' in query and ('trend' in query or 'popular' in query):
            suggestions = get_product_suggestions()
            product_names = ", ".join([s['Name'] for s in suggestions[:3]])
            response = f"Popular suggestions: {product_names}"
        elif 'restock' in query or 'low stock' in query:
            predictions = generate_restock_predictions()
            if predictions:
                response = "You might want to restock: " + ", ".join([p['product'] for p in predictions])
            else:
                response = "Your stock levels look good right now."
        elif 'combo' in query or 'deal' in query:
            combos = generate_combo_suggestions()
            if combos:
                response = "Suggested combos: " + "; ".join([f"{c['products']} ({c['discount']})" for c in combos])
            else:
                response = "No combo suggestions available right now."
        else:
            response = "I can help you track orders, suggest products, check restock needs, or find combo deals. Please ask specifically."
        
        return render_template('ai_assistant.html', query=query, response=response)
    
    return render_template('ai_assistant.html')

@app.route('/profile')
def profile():
    if 'email' not in session:
        return redirect(url_for('login'))
    
    try:
        df = pd.read_excel(USERS_FILE)
        user = df[df['Email'] == session['email']].iloc[0].to_dict()
        
        rewards_df = pd.read_excel(REWARDS_FILE)
        rewards = rewards_df.iloc[0].to_dict() if not rewards_df.empty else None
        
        money_df = pd.read_excel(MONEY_SPENT_FILE)
        total_spent = float(money_df['Amount'].sum()) if not money_df.empty else 0.0
        
        return render_template('profile.html', 
                             user=user, 
                             rewards=rewards,
                             total_spent=total_spent)
    except Exception as e:
        print(f"Error loading profile data: {e}")
        return render_template('profile.html', error="Could not load profile data")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)